# coding: utf-8
import csv
from io import BytesIO

from django.urls import reverse
from django.core.files.storage import get_storage_class, FileSystemStorage
from openpyxl import load_workbook

from onadata.apps.viewer.models.export import Export
from onadata.apps.viewer.views import export_download
from onadata.libs.utils.export_tools import generate_export
from onadata.libs.utils.user_auth import http_auth_string
from .test_base import TestBase


class TestFormExports(TestBase):

    def setUp(self):
        TestBase.setUp(self)
        self._create_user_and_login()
        self._publish_transportation_form_and_submit_instance()

    def _num_rows(self, content, export_format):
        def xls_rows(f):
            wb = load_workbook(BytesIO(f))
            return wb[wb.sheetnames[0]].max_row

        def csv_rows(f):
            return len([line for line in csv.reader(f.decode().strip().split('\n'))])
        num_rows_fn = {
            'xls': xls_rows,
            'csv': csv_rows,
        }
        return num_rows_fn[export_format](content)

    def test_allow_export_download_for_basic_auth(self):
        extra = {
            'HTTP_AUTHORIZATION': http_auth_string(self.login_username,
                                                   self.login_password)
        }
        # create export
        export = generate_export(Export.CSV_EXPORT, 'csv', self.user.username,
                                 self.xform.id_string)
        self.assertTrue(isinstance(export, Export))
        url = reverse(export_download, kwargs={
            'username': self.user.username,
            'id_string': self.xform.id_string,
            'export_type': export.export_type,
            'filename': export.filename
        })
        response = self.anon.get(url, **extra)
        default_storage = get_storage_class()()
        if not isinstance(default_storage, FileSystemStorage):
            self.assertEqual(response.status_code, 302)
        else:
            self.assertEqual(response.status_code, 200)
